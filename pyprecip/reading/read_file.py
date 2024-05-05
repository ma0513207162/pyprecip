import os, csv
from openpyxl import load_workbook
from ..utits.except_ import RaiseException as exc 
from ..utits.warn_ import RaiseWarn as warn 
from ..utits.sundries import check_param_type


def __file_exists(func):
    def wrapper(*args, **kwargs):     
        if not args and not kwargs:
            exc.raise_exception("The file path is required.", TypeError) 
        return func(*args, **kwargs)
    return wrapper 



@__file_exists    
def read_csv(path: str, row_indices: (tuple|list) = (), column_indices: (tuple|list) = ()): 
    """     
    Reads data for specified rows and columns from a CSV file.

    Parameters:
     - path: indicates the path of the CSV file
     - row_indices: Specifies the read row range (list length 2) or row index (tuple)
     - column_indices: specifies the read column range (list length 2) or column index (tuple)

    Returns:
     - Dictionary. The key indicates the file name and the value indicates the read data
    """

    check_param_type(path, str, "path");
    check_param_type(row_indices, (tuple|list), "row_indices"); 
    check_param_type(column_indices, (tuple|list), "column_indices");

    read_csv_result: dict = {}; 
    with open(path, "r", encoding="gbk") as csv_file:
        reader_csv = list(csv.reader(csv_file)) 

    # 指定行范围  
    if isinstance(row_indices, list) and row_indices != []:
        if len(row_indices) == 2:
            start, end = row_indices[0], row_indices[1]
            reader_csv = reader_csv[start-1: end]
        else:
            warn.raise_warning("The row_indices parameter must contain only two elements, otherwise it is invalid.") 
    
    # 指定行索引 
    if isinstance(row_indices, tuple) and row_indices != ():
        row_idx_list = []
        for idx in row_indices:
            if idx >= 1 and idx <= len(reader_csv):
                row_idx_list.append(reader_csv[idx-1]) 
            else:
                exc.raise_exception("The index must be greater than 0 and less than the sequence length.", IndexError)
        reader_csv = row_idx_list; 

    # list 类型指定行范围
    reader_csv = list(zip(*reader_csv)) 
    if isinstance(column_indices, list) and column_indices != []:
        if len(column_indices) == 2:
            start, end = column_indices[0], column_indices[1]; 
            reader_csv = reader_csv[start-1: end]
        else:
            warn.raise_warning("The column_indices parameter must contain only two elements, otherwise it is invalid.") 
    
    # tuple 类型指定列索引 
    if isinstance(column_indices, tuple) and column_indices != ():
        col_idx_list = [] 
        for idx in column_indices:
            if idx >= 1 and idx <= len(reader_csv):
                col_idx_list.append(reader_csv[idx-1]); 
            else:
                exc.raise_exception("The index must be greater than 0 and less than the sequence length.", IndexError)
        reader_csv = col_idx_list;
    
    # 封装 dict 对象
    file_name = os.path.splitext(os.path.basename(path))[0]; 
    read_csv_result[file_name] = reader_csv; 
    return read_csv_result;



@__file_exists    
def read_excel(path: str, sheet_names: tuple = (), row_indices: (tuple|list) = (),
            column_indices: (tuple|list) = ()) -> dict: 
    workbook = load_workbook(filename = path, data_only = True) 

    # Gets the specified worksheet 
    sheet_list = []
    if sheet_names != ():
        for sheet_name in sheet_names:
            sheet_list.append(workbook[sheet_name])
    else:
        sheet_list.append(workbook.active) 

    read_excel_result: dict = {}; 
    # 遍历工作簿 sheet_list
    for sheet in sheet_list:
        sheet_iter_rows: list = list(sheet.iter_rows(values_only = True)) 

        # 指定行范围 
        if isinstance(row_indices, list) and row_indices != []:
            if len(row_indices) == 2:
                start, end = row_indices[0], row_indices[1] 
                sheet_iter_rows = sheet_iter_rows[start-1: end]
            else:
                warn.raise_warning("The row_indices parameter must contain only two elements, otherwise it is invalid.") 

        # 指定行索引 
        if isinstance(row_indices, tuple) and row_indices != ():
            temp_iter_rows = []
            for idx in row_indices:
                if idx > 0 and idx <= len(sheet_iter_rows):
                    temp_iter_rows.append(sheet_iter_rows[idx-1]) 
                else:
                    exc.raise_exception("The index must be greater than 0 and less than the sequence length.", IndexError)
            sheet_iter_rows = temp_iter_rows   
        
        # list 类型指定行范围
        sheet_iter_rows = list(zip(*sheet_iter_rows)) 
        if isinstance(column_indices, list) and column_indices != []:
            if len(column_indices) == 2:
                start, end = column_indices[0], column_indices[1]; 
                sheet_iter_rows = sheet_iter_rows[start-1: end]  
            else:
                warn.raise_warning("The column_indices parameter must contain only two elements, otherwise it is invalid.")   

        # tuple 类型指定列索引 
        if isinstance(column_indices, tuple) and column_indices != ():
            col_idx_list = [] 
            for idx in column_indices:
                if idx >= 1 and idx <= len(sheet_iter_rows):
                    col_idx_list.append(sheet_iter_rows[idx-1]); 
                else:
                    exc.raise_exception("The index must be greater than 0 and less than the sequence length.", IndexError)
            sheet_iter_rows = col_idx_list; 


        read_excel_result[sheet.title] = sheet_iter_rows; 


    return read_excel_result; 
 



# 以主进程的方式运行 
if __name__ == "__main__": 
    path = "./static/test_data.csv"
    reader_csv = read_csv(path=path, row_indices=[1,22,3,3]);

    print(reader_csv)
